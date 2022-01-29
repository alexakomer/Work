import os
import pandas as pd
from openpyxl import load_workbook


def compare(outputFile, compareFile, goldenFile, indexVal, criticalKeys):

    recipeName, extension = goldenFile.split(".")

    fileName = goldenFile
    while not os.path.isfile(fileName):
        print("The " + fileName + " file isn't accessible.")
        exit()

    goldenFile = open(fileName, "r")

    # open secondary comparison file

    fileName2 = compareFile
    while not os.path.isfile(fileName2):
        print("The " + fileName2 + " file isn't accessible.")
        exit()

    compareFile = open(fileName2, "r")

    # populate golden file's dictionary

    golden_dictionary = {}

    for line in goldenFile:
        key, value = line.split("=")
        golden_dictionary[key] = value

    for key in golden_dictionary:
        golden_dictionary[key] = golden_dictionary[key].replace('\n', "").strip()

    # populate comparison file's dictionary

    compare_dictionary = {}

    for line in compareFile:
        key, value = line.split("=")
        compare_dictionary[key] = value

    for key in compare_dictionary:
        compare_dictionary[key] = compare_dictionary[key].replace('\n', "").strip()

    # ensure both files are closed

    goldenFile.close()
    compareFile.close()

    # compare each value and count critical values

    differenceCount = 0
    criticalCount = 0
    index = 0
    for keys in golden_dictionary:
        if golden_dictionary[keys] != compare_dictionary[keys] and criticalKeys[index]:
            criticalCount += 1
        if golden_dictionary[keys] != compare_dictionary[keys]:
            differenceCount += 1
        index += 1

    # create difference and filepath list

    # specify absolute file path? currently relative.
    # if you want absolute, use a concatenation of os.getcwd() and fileName2

    differenceList = [fileName2, criticalCount, differenceCount]

    # create difference and filepath dataframe

    differenceFrame = pd.DataFrame(differenceList)

    # create data dataframes

    golden_data = pd.read_csv(fileName, names = ["Parameters", "Values"], delimiter = "=", skiprows=0)
    compare_data = pd.read_csv(fileName2, names = ["Parameters", "Values"], delimiter = "=", skiprows=0)

    # load excel file

    wb = load_workbook(outputFile)
    # print(wb.sheetnames)

    writer = pd.ExcelWriter(outputFile, engine='openpyxl')

    writer.book = wb
    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

    # write in critical keys and titles on far left

    criticalFrame = pd.DataFrame(criticalKeys, columns=['Critical Keys'])

    criticalFrame.to_excel(writer, sheet_name=recipeName, columns=['Critical Keys'], index=False, startcol=0,startrow=3)

    titleFrame = pd.DataFrame(["Filepath:", "Critical Differences:", "Total Differences:"])

    titleFrame.to_excel(writer, sheet_name=recipeName, header=None, index=False, startcol=0, startrow=0)

    # write data to csv

    golden_data.to_excel(writer, sheet_name=recipeName, header="Parameters", index=False, startcol=1,startrow=3) # will be done once
    compare_data.to_excel(writer, sheet_name=recipeName, columns= ["Values"], index=False, startcol=1 + indexVal,startrow=3) # will be iterated (add counter to col)

    # write filepath and differences above each column of values

    differenceFrame.to_excel(writer, sheet_name=recipeName, header=None, index=False, startcol=1 + indexVal,startrow=0) # will be iterated (add counter to col)

    # remember to save

    wb.save(outputFile)


outFile = 'Comparisons.xlsx' # Excel file with this name must have already been created. Be careful, this will overwrite existing information


# version 1 -----------------------------------------------------------------------------------------------
# recipeFrame = pd.read_excel("DirectoryList.xlsx")
# for dirpath, dirnames, filenames in os.walk("."):
#     for filename in [f for f in filenames if f.endswith(inFile)]:
#         counter += 1
#         # print (counter, " ", os.path.join(dirpath, filename))
#         compare(outFile, os.path.join(dirpath, filename), inFile, counter)
# ---------------------------------------------------------------------------------------------------------


# version 2 -----------------------------------------------------------------------------------------------
xl = pd.ExcelFile("DirectoryList.xlsx")

i = 0

recipeList = []

criticalKeyList = []

for sheet in xl.sheet_names:
    if sheet != 'Sheet1':
        toggleFrame = xl.parse(sheet, nrows=1, header= None, names= ['Toggle', 'Recipe'])
        dataFrame = xl.parse(sheet, skiprows=3)
        if toggleFrame.loc[0].at['Toggle']:
            recipeList.append(toggleFrame.loc[0].at['Recipe'])
            temp = []
            k = 0
            while k < (len(list(dataFrame.Parameters))):
                temp.append(dataFrame.loc[k].at['Critical Keys'])
                k += 1
            criticalKeyList.append(temp)
            i += 1

i = 0
while i < len(recipeList):
    inFile = recipeList[i]
    counter = 0
    for dirpath, dirnames, filenames in os.walk("."):
            for filename in [f for f in filenames if f.endswith(inFile + ".rep")]: # or f.endswith(inFile + ".txt") or f.endswith(inFile + ".ini")]:
                counter += 1
                # print (counter, " ", os.path.join(dirpath, filename))
                compare(outFile, os.path.join(dirpath, filename), inFile + ".rep", counter, criticalKeyList[i])
#                 # compare(outFile, os.path.join(dirpath, filename), inFile + ".txt", counter)
#                 # compare(outFile, os.path.join(dirpath, filename), inFile + ".ini", counter)
    i += 1
# ---------------------------------------------------------------------------------------------------------

# i = 0
# while i < len(list(recipeFrame.Recipes)):
#     if recipeFrame.loc[i].at["Include?"]:
#         inFile = recipeFrame.loc[i].at["Recipes"]
#         counter = 0
#         for dirpath, dirnames, filenames in os.walk("."):
#             for filename in [f for f in filenames if f.endswith(inFile + ".rep")]: # or f.endswith(inFile + ".txt") or f.endswith(inFile + ".ini")]:
#                 counter += 1
#                 # print (counter, " ", os.path.join(dirpath, filename))
#                 compare(outFile, os.path.join(dirpath, filename), inFile + ".rep", counter)
#                 # compare(outFile, os.path.join(dirpath, filename), inFile + ".txt", counter)
#                 # compare(outFile, os.path.join(dirpath, filename), inFile + ".ini", counter)
#     i += 1
# ---------------------------------------------------------------------------------------------------------


# if I want to attempt applying conditional formatting:
# first, read the excel sheet back into a full dataframe using pandas
# and then use df.style.apply to possibly create formatting
# parse from each sheet
