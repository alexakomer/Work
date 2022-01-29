import tkinter as tk
from tkinter import *
from tkinter import filedialog
import os.path

keepGoing = True

while(keepGoing):

    # Dialog Box ---------------------------------------------------------------------------------

    root = Tk()
    root.title("Test Dialog Box")
    root.geometry("300x130")


    def focus_next_window(event):  # Create tab controls
        event.widget.tk_focusNext().focus()
        return ("break")


    def enter(event):  # Create function to handle pressing enter
        update()
        return ("break")


    def update():
        folder_name_temp = browseFiles()
        global params
        params = [t1.get('1.0', END).strip(), t2.get('1.0', END).strip(), t3.get('1.0', END).strip(), folder_name_temp]
        root.destroy()


    # focus_next_window()

    l1 = tk.Label(root, text="Customer Name:")
    l1.grid(row=1, column=1)

    t1 = tk.Text(root, height=1, width=20, borderwidth=5)
    t1.grid(row=1, column=2)
    t1.focus()
    t1.bind("<Tab>", focus_next_window)
    t1.bind('<Return>', enter)

    l2 = tk.Label(root, text="Customer Site:")
    l2.grid(row=2, column=1)

    t2 = tk.Text(root, height=1, width=20, borderwidth=5)
    t2.grid(row=2, column=2)
    t2.bind("<Tab>", focus_next_window)
    t2.bind('<Return>', enter)

    l3 = tk.Label(root, text="Customer Tool ID:")
    l3.grid(row=3, column=1)

    t3 = tk.Text(root, height=1, width=20, borderwidth=5)
    t3.grid(row=3, column=2)
    t3.bind("<Tab>", focus_next_window)
    t3.bind('<Return>', enter)

    b1 = tk.Button(root, text="Ok", command=lambda: update())
    b1.grid(row=5, column=2)


    def browseFiles():
        folder_name_temp = tk.filedialog.askdirectory()
        return folder_name_temp


    root.mainloop()

    customer_name, customer_site, customer_tool, folder_name_temp = params

    # print(folder_name_temp)

    # -----------------------------------------------------------------------------------------------

    # done with user input

    import os

    # save old directory path
    og_path = os.getcwd()

    # work in the requested directory

    os.chdir(folder_name_temp)

    recipeList = []

    # Determine Recipe List ---------------------------------------------------------------------------------

    # walk through directory / subdirectories and determine what recipes are here

    for dirpath, dirnames, filenames in os.walk("."):
        for filename in [f for f in filenames]:
            recipeList.append(os.path.join(os.getcwd(), dirpath[2:], filename))

    toplevel_dir = os.listdir(os.getcwd())

    top_folder_arr = folder_name_temp.split("/")
    top_folder = top_folder_arr[len(top_folder_arr) - 1]

    # print(top_folder)

    # -----------------------------------------------------------------------------------------------


    # Begin Excel IO ---------------------------------------------------------------------------------


    # read from and write to excel

    import pandas as pd
    from openpyxl import load_workbook
    from os.path import exists

    os.chdir(og_path) # make sure to set this correctly

    macroBool = exists("MFC-EV.xlsm")

    if macroBool: # handles xlsm files
        workbook_name = "MFC-EV.xlsm"
        wb = load_workbook(workbook_name, data_only=True, keep_vba=True)

    else:         # handles xlsx files
        workbook_name = "MFC-EV.xlsx"
        wb = load_workbook(workbook_name, data_only=True)

    writer = pd.ExcelWriter(workbook_name, engine='openpyxl')

    writer.book = wb
    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

    from datetime import date, datetime
    import time

    total_diff = 0
    total_crit = 0
    critical_recipe_list = []

    def traverse(folder_name, tabName): # will be traversed iteratively by a loop following the function, similar to comparator.py

        global total_diff
        global total_crit
        global critical_recipe_list

        os.chdir(folder_name)

        directory_name_arr = folder_name.split("/")
        directory_name = directory_name_arr[len(directory_name_arr) - 1]


        if directory_name not in wb.sheetnames:
            directory_name = tabName

        criticals = []
        dates = []
        times = []
        sizes = []
        names = []
        tabs = []

        num_crits = wb[directory_name].cell(9,1).value

        for row in range(11, wb[directory_name].max_row + 1):
            criticals.append(wb[directory_name].cell(row, 1).value)
            dates.append(wb[directory_name].cell(row, 2).value)
            times.append(wb[directory_name].cell(row, 3).value)
            times.append(wb[directory_name].cell(row, 3).value)
            sizes.append(wb[directory_name].cell(row, 4).value)
            names.append(wb[directory_name].cell(row, 5).value)
            tabs.append(wb[directory_name].cell(row, 9).value)

        file_list = []
        time_list = []
        size_list = []

        # Create lists with file names in directory ---------------------------------------------------------------------------------

        # fill in file name, time, and size (handle directory size as <DIR>)
        for i in os.listdir(folder_name):
            a = os.stat(os.path.join(folder_name, i))
            file_list.append(i) # file
            time_list.append(time.ctime(os.stat(os.path.join(folder_name, i)).st_mtime)) # most recent modification
            if (os.path.isfile(os.path.join(folder_name, i))):
                size_list.append(os.path.getsize(os.path.join(folder_name, i)))
            else:
                size_list.append("<DIR>")

        walk = 0
        # print(directory_name)
        while not (wb[directory_name].cell(column=10 + walk, row=2).value == None):
            walk += 3

        customer_name_frame = pd.DataFrame([customer_name])
        customer_name_frame.to_excel(writer, sheet_name=directory_name, header=None, index=False, startcol=9 + walk, startrow=1)

        customer_site_frame = pd.DataFrame([customer_site])
        customer_site_frame.to_excel(writer, sheet_name=directory_name, header=None, index=False, startcol=9 + walk, startrow=2)

        customer_tool_frame = pd.DataFrame([customer_tool])
        customer_tool_frame.to_excel(writer, sheet_name=directory_name, header=None, index=False, startcol=9 + walk, startrow=3)

        time_frame = pd.DataFrame([datetime.now()])
        time_frame.to_excel(writer, sheet_name=directory_name, header=None, index=False, startcol=9 + walk, startrow=4)

        file_frame = pd.DataFrame([directory_name])
        file_frame.to_excel(writer, sheet_name=directory_name, header=None, index=False, startcol=9 + walk, startrow=5)

        folder_frame = pd.DataFrame([folder_name])
        folder_frame.to_excel(writer, sheet_name=directory_name, header=None, index=False, startcol=9 + walk, startrow=6)


        # leftovers handle files that exist in structure but not in preset excel file ------------------------------------------------
        leftover_time = []
        leftover_size = []
        leftover_file = []

        for i in range(0, len(file_list)):
            found = False
            for j in range(0, len(names)):
                if(str(file_list[i]) == str(names[j])):
                    found = True
                    file_frame = pd.DataFrame([[time_list[i], size_list[i], file_list[i]]], columns=["Date", "Size", "Name"])
                    file_frame.to_excel(writer, sheet_name=directory_name, header=None, index=False, startcol=9 + walk,startrow= 10 + j)

            if(not found):
                leftover_time.append(time_list[i])
                leftover_size.append(size_list[i])
                leftover_file.append(file_list[i])

        # ----------------------------------------------------------------------------------------------------------------------------


        # handle differences in largest directory tab in excel
        differences = 0
        critical_differences = 0

        for row in range(11, 11 + int(num_crits)):
            if (wb[directory_name].cell(row, 12).value == None): #handles blank lines
                differences += 1
                if (criticals[row - 11]):
                    critical_differences += 1

        total_diff += differences
        total_crit += critical_differences

        if (critical_differences >= 1):
            critical_recipe_list.append(directory_name)

        # Creating new critical tab

        critical_difference_frame = pd.DataFrame([critical_differences])
        critical_difference_frame.to_excel(writer, sheet_name=directory_name, header=None, index=False, startcol=9 + walk, startrow=7)

        difference_frame = pd.DataFrame([differences])
        difference_frame.to_excel(writer, sheet_name=directory_name, header=None, index=False, startcol=9 + walk, startrow=8)

        leftover_frame = pd.DataFrame(
            {'Date': leftover_time,
             'Size': leftover_size,
             'Name': leftover_file
            })

        leftover_frame.to_excel(writer, sheet_name=directory_name, header=None, index=False, startcol=9 + walk,startrow= 13 + len(file_list)) # will be done once

        from shutil import copyfile

        for i in range(len(criticals)):
            if(criticals[i]):
                if not os.path.exists(folder_name + "/" + str(names[i])):
                    print("file " + str(names[i]) + " doesn't exist in the " + directory_name + " tab.")
                    continue
                else:
                    if(sizes[i] != "<DIR>"):

                        filename, extension = names[i].split(".")
                        if (extension == "ini"):

                            dst = "temp_" + filename +".txt"

                            copyfile(folder_name + "/" + names[i], folder_name + "/" + dst)
                            fileObject = open(folder_name + "/" + dst, "r")
                            compareDataString = fileObject.read()
                            fileObject.close()
                            compareArray = compareDataString.splitlines()
                            compareData = pd.DataFrame(compareArray)

                            os.remove(folder_name + "/" + dst)

                        else:
                            compareData = pd.read_csv(folder_name + "/" + names[i], header=None)

                        # CHECK IF DATA EXISTS AND WALK OVER APPROPRIATE NUM OF COLUMNS

                        check = 0
                        while not (wb[tabs[i]].cell(row=2, column=10 + check).value == None):
                            check += 1

                        compareData.to_excel(writer, sheet_name=tabs[i], header=None, index=False, startcol=9 + check, startrow=10)

                        customer_name_frame = pd.DataFrame([customer_name])
                        customer_name_frame.to_excel(writer, sheet_name=tabs[i], header=None, index=False, startcol=9 + check, startrow=1)

                        customer_site_frame = pd.DataFrame([customer_site])
                        customer_site_frame.to_excel(writer, sheet_name=tabs[i], header=None, index=False, startcol=9 + check, startrow=2)

                        customer_tool_frame = pd.DataFrame([customer_tool])
                        customer_tool_frame.to_excel(writer, sheet_name=tabs[i], header=None, index=False, startcol=9 + check, startrow=3)

                        time_frame = pd.DataFrame([datetime.now()])
                        time_frame.to_excel(writer, sheet_name=tabs[i], header=None, index=False, startcol=9 + check, startrow=4)

                        file_frame = pd.DataFrame([names[i]])
                        file_frame.to_excel(writer, sheet_name=tabs[i], header=None, index=False, startcol=9 + check, startrow=5)

                        folder_frame = pd.DataFrame([folder_name])
                        folder_frame.to_excel(writer, sheet_name=tabs[i], header=None, index=False, startcol=9 + check, startrow=6)

                        differences = 0
                        critical_differences = 0

                        new_criticals = []

                        for row in range(11, wb[tabs[i]].max_row + 1):
                            new_criticals.append(wb[tabs[i]].cell(row, 1).value)

                        for row in range(11, 11 + len(new_criticals)):
                            if (wb[tabs[i]].cell(row, 10 + check).value != wb[tabs[i]].cell(row, 2).value):
                                differences += 1
                                if ((wb[tabs[i]].cell(row, 10 + check).value == "") and (wb[tabs[i]].cell(row, 2).value == None)):
                                    differences -= 1 # to avoid counting blank line differences
                                if (new_criticals[row - 11]):
                                    critical_differences += 1
                                    differences -= 1 # to avoid double counting differences

                        total_crit += critical_differences
                        total_diff += differences

                        if (critical_differences >= 1):
                            critical_recipe_list.append(tabs[i])

                        critical_difference_frame = pd.DataFrame([critical_differences])
                        critical_difference_frame.to_excel(writer, sheet_name=tabs[i], header=None, index=False, startcol=9 + check, startrow=7)

                        difference_frame = pd.DataFrame([differences])
                        difference_frame.to_excel(writer, sheet_name=tabs[i], header=None, index=False, startcol=9 + check, startrow=8)

                    else:
                        new_folder = folder_name + "/" + str(names[i])
                        if (os.path.exists(new_folder)):
                            traverse(new_folder, tabs[i])

    # run recursive function using "eagleview" as top level folder, for now.

    traverse(folder_name_temp, "eagleview")

    customer_data = [[customer_name, customer_site, customer_tool]]

    today = date.today()
    today = today.strftime("%m/%d/%Y")
    customer_data[0].append(today)
    now = datetime.now()
    now = now.strftime("%H:%M")
    customer_data[0].append(now)

    customer_data[0].append(total_crit)
    customer_data[0].append(total_diff)

    # append number of critical, noncritical, and unique files -----------------------------------------------------------------------

    customer_frame = pd.DataFrame(customer_data, columns=["Customer", "Site", "Tool", "Capture Date", "Capture Time", "# Critical", "# Non-Critical"])

    customer_walk = 0
    while not (wb["Summary_Sheet"].cell(column=4, row=9 + customer_walk).value == None):
        customer_walk += 1

    customer_frame.to_excel(writer, sheet_name="Summary_Sheet", header=None, index=False, startcol=0,startrow=8 + customer_walk) # will be done once

    recipe_walk = 0
    while not (wb["Critical_Differences"].cell(column=2 + recipe_walk, row=1).value == None):
        recipe_walk += 1

    # finish by writing data to Critical_Differences tab

    customer_name_frame = pd.DataFrame([customer_name])
    customer_name_frame.to_excel(writer, sheet_name="Critical_Differences", header=None, index=False, startcol=1 + recipe_walk, startrow=0)

    customer_site_frame = pd.DataFrame([customer_site])
    customer_site_frame.to_excel(writer, sheet_name="Critical_Differences", header=None, index=False, startcol=1 + recipe_walk, startrow=1)

    customer_tool_frame = pd.DataFrame([customer_tool])
    customer_tool_frame.to_excel(writer, sheet_name="Critical_Differences", header=None, index=False, startcol=1 + recipe_walk, startrow=2)

    time_frame = pd.DataFrame([datetime.now()])
    time_frame.to_excel(writer, sheet_name="Critical_Differences", header=None, index=False, startcol=1 + recipe_walk, startrow=3)

    critical_difference_frame = pd.DataFrame([total_crit])
    critical_difference_frame.to_excel(writer, sheet_name="Critical_Differences", header=None, index=False, startcol=1 + recipe_walk, startrow=4)

    difference_frame = pd.DataFrame([total_diff])
    difference_frame.to_excel(writer, sheet_name="Critical_Differences", header=None, index=False, startcol=1 + recipe_walk, startrow=5)


    critical_recipe_frame = pd.DataFrame([critical_recipe_list])
    critical_recipe_frame = critical_recipe_frame.T
    critical_recipe_frame.to_excel(writer, sheet_name = "Critical_Differences", header = None, index = False, startcol=1 + recipe_walk, startrow=6)

    print("Code finished without errors.")

    os.chdir(og_path)

    wb.save(workbook_name)

    response = input(
        "1.) Press \"1\" to run this script again.\n2.) Press \"2\" to close this script.\n3.) Press \"3\" to close this script and open the Excel file\n")

    if (response == "1"):
        keepGoing = True

    if (response == "2"):
        keepGoing = False
        # exit()
        quit()

    if (response == "3"):
        launch_str = "start EXCEL.EXE " + workbook_name
        os.system(launch_str)
        keepGoing = False
        # exit()
        quit()



    # Possible later implementations:
    # - Try recreating this script to be called from Excel using VBA (most polished)
    # - Use xlwings to rewrite entire script for xlsm cases (should account for VBA on its own)
    # - Use example code using idea that xlsm and xlsx files are zip files (though encountering error with code below)
    # - The given macros are simple - try to recreate them at the end of the Python script
    # - Recreate VBA Macros as Python scripts to be run as executables (least polished)